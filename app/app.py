from flask import session
import threading
import time

# In-memory progress store (for demo; use Redis or DB for production)
progress_store = {}

import os
from flask import Flask, request, send_from_directory, jsonify, render_template_string
from werkzeug.utils import secure_filename
import pandas as pd
import fitz  # PyMuPDF
import csv
import re
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import pytesseract
from PIL import Image
import io
import uuid

# Use absolute paths for directories to avoid issues in cloud environments
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
EXPORTS_FOLDER = os.path.join(BASE_DIR, 'exports')
ALLOWED_EXTENSIONS = {'pdf', 'csv', 'xlsx'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['EXPORTS_FOLDER'] = EXPORTS_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_with_ocr(page):
    text = page.get_text()
    if text.strip():
        return text
    # Fallback to OCR if no text found
    pix = page.get_pixmap(dpi=150)
    img = Image.open(io.BytesIO(pix.tobytes("png")))
    text = pytesseract.image_to_string(img)
    img.close()
    return text

def validate_pdf(pdf_path, export_dir, progress_key=None, result_key=None):
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    csv_path = os.path.join(export_dir, f"{base_name}_validation_summary.csv")
    excel_path = os.path.join(export_dir, f"{base_name}_validation_summary.xlsx")
    dashboard_path = os.path.join(export_dir, f"{base_name}_dashboard.png")

    REQUIRED_FIELDS = [
        "Customer Name", "Customer P.O. Number", "Customer Part Number",
        "Customer Part Number Revision", "OEM Part Number", "OEM Lot Number",
        "OEM Date Code", "OEM Cage Code", "AEM Part Number", "AEM Lot Number",
        "AEM Date Code", "AEM Cage Code", "Customer Quality Clauses",
        "FAI Form 3", "Solderability Test Report", "DPA", "Visual Inspection Record",
        "Shipment Quantity", "Reel Labels", "Certificate of Conformance", "Route Sheet",
        "Part Number", "Lot Number", "Date", "Resistance", "Dimension", "Test Result"
    ]

    NUMERICAL_RANGES = {
        "Resistance": (95, 105),
        "Dimension": (0.9, 1.1)
    }

    anomalies = []
    critical_issues = []
    field_presence = defaultdict(int)
    all_fields = []

    def extract_fields(text):
        fields = {}
        for field in REQUIRED_FIELDS:
            pattern = rf"{field}[:\s]*([^\n]+)"
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                fields[field] = match.group(1).strip()
        return fields

    def validate_numerical(field, value):
        try:
            val = float(re.findall(r"[\d.]+", value)[0])
            min_val, max_val = NUMERICAL_RANGES[field]
            return min_val <= val <= max_val
        except:
            return False

    def check_consistency(field_name):
        values = [fields.get(field_name) for fields in all_fields if field_name in fields]
        return len(set(values)) == 1

    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    for page_num in range(total_pages):
        page = doc.load_page(page_num)
        text = extract_text_with_ocr(page)
        fields = extract_fields(text)
        all_fields.append(fields)

        for field in REQUIRED_FIELDS:
            if field not in fields:
                anomalies.append([page_num + 1, field, "Missing"])
            else:
                field_presence[field] += 1

        for field in NUMERICAL_RANGES:
            if field in fields and not validate_numerical(field, fields[field]):
                anomalies.append([page_num + 1, field, f"Out of range: {fields[field]}"])
                critical_issues.append([page_num + 1, field, fields[field]])

        # Update progress
        if progress_key:
            progress_store[progress_key]['percent'] = int(((page_num + 1) / total_pages) * 100)
        print(f"Processed page {page_num+1}/{total_pages}")

    for field in ["Part Number", "Lot Number", "Date"]:
        if not check_consistency(field):
            anomalies.append(["All Pages", field, "Inconsistent values"])
            critical_issues.append(["All Pages", field, "Inconsistent values"])

    with open(csv_path, "w", newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Page", "Field", "Issue"])
        writer.writerows(anomalies)

    wb = Workbook()
    ws = wb.active
    ws.title = "QA Anomalies"

    headers = ["Page", "Field", "Issue"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(anomalies, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    table_ref = f"A1:C{len(anomalies)+1}"
    table = Table(displayName="AnomalyTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(excel_path)

    plt.figure(figsize=(12, 6))
    plt.bar(field_presence.keys(), field_presence.values(), color='skyblue')
    plt.title("Field Presence Across PDF Pages")
    plt.xlabel("Field Name")
    plt.ylabel("Number of Pages Present")
    plt.xticks(rotation=90)
    plt.tight_layout()
    plt.savefig(dashboard_path)

    # Save result in progress_store for robust retrieval
    if progress_key and result_key:
        progress_store[progress_key]['percent'] = 100
        progress_store[progress_key]['csv_filename'] = os.path.basename(csv_path)
        progress_store[progress_key]['done'] = True

    print(f"Validation complete. CSV saved at {csv_path}")
    return csv_path, excel_path, dashboard_path, len(anomalies), len(critical_issues)

def validate_file(filepath, progress_key=None, result_key=None):
    # If PDF, run PDF validation, else fallback to dummy
    if filepath.lower().endswith('.pdf'):
        df = None
        csv_path, excel_path, dashboard_path, anomaly_count, critical_count = validate_pdf(filepath, EXPORTS_FOLDER, progress_key, result_key)
        df = pd.read_csv(csv_path)
        print(f"validate_file: CSV generated at {csv_path}")
        return df, os.path.basename(csv_path)
    else:
        data = {'filename': [os.path.basename(filepath)], 'status': ['validated']}
        df = pd.DataFrame(data)
        csv_filename = os.path.splitext(os.path.basename(filepath))[0] + '.csv'
        csv_path = os.path.join(EXPORTS_FOLDER, csv_filename)
        df.to_csv(csv_path, index=False)
        if progress_key and result_key:
            progress_store[progress_key]['percent'] = 100
            progress_store[progress_key]['csv_filename'] = csv_filename
            progress_store[progress_key]['done'] = True
        print(f"validate_file: Non-PDF CSV generated at {csv_path}")
        return df, csv_filename

def export_to_csv(df, csv_path):
    df.to_csv(csv_path, index=False)

@app.route('/', methods=['GET'])
def index():
    return render_template_string('''
    <h2>Upload file for validation</h2>
    <p>1. Select a file.<br>
    2. Click <b>Upload and Validate</b>.<br>
    3. Wait for both progress bars to reach 100%.<br>
    4. When validation is complete, click the <b>Download CSV</b> link.</p>
    <form id="upload-form" method="post" action="/api/validate" enctype="multipart/form-data">
      <input type="file" name="file" id="file-input">
      <input type="submit" value="Upload and Validate">
    </form>
    <div style="margin-top:20px;">
      <div>Upload Progress: <span id="upload-percent">0%</span></div>
      <div id="upload-progress-bar" style="width: 100%; background: #eee; height: 20px;">
        <div id="upload-progress" style="background: #2196f3; width: 0%; height: 100%;"></div>
      </div>
    </div>
    <div style="margin-top:20px;">
      <div>Validation Progress: <span id="progress-percent">0%</span></div>
      <div id="progress-bar" style="width: 100%; background: #eee; height: 20px;">
        <div id="progress" style="background: #4caf50; width: 0%; height: 100%;"></div>
      </div>
    </div>
    <div id="download-link" style="margin-top:20px;"></div>
    <script>
    document.getElementById('upload-form').onsubmit = async function(e) {
      e.preventDefault();
      const formData = new FormData(this);
      let uploadProgressBar = document.getElementById('upload-progress');
      let uploadPercentText = document.getElementById('upload-percent');
      let progressBar = document.getElementById('progress');
      let progressPercentText = document.getElementById('progress-percent');
      uploadProgressBar.style.width = '0%';
      uploadPercentText.innerText = '0%';
      progressBar.style.width = '0%';
      progressPercentText.innerText = '0%';
      document.getElementById('download-link').innerHTML = '';

      // AJAX upload with progress
      const xhr = new XMLHttpRequest();
      xhr.open('POST', '/api/validate', true);

      xhr.upload.onprogress = function(e) {
        if (e.lengthComputable) {
          let percent = Math.round((e.loaded / e.total) * 100);
          uploadProgressBar.style.width = percent + '%';
          uploadPercentText.innerText = percent + '%';
        }
      };

      xhr.onreadystatechange = async function() {
        if (xhr.readyState === XMLHttpRequest.DONE) {
          if (xhr.status === 200) {
            uploadProgressBar.style.width = '100%';
            uploadPercentText.innerText = '100%';
            const data = JSON.parse(xhr.responseText);
            if (!data.progressKey) {
              document.getElementById('download-link').innerText = 'Validation failed.';
              return;
            }
            // Poll for validation progress
            let percent = 0;
            let csvFilename = '';
            while (percent < 100) {
              const progRes = await fetch(`/api/progress/${data.progressKey}`);
              const progData = await progRes.json();
              percent = progData.percent;
              progressBar.style.width = percent + '%';
              progressPercentText.innerText = percent + '%';
              if (progData.done && progData.csv_filename) {
                csvFilename = progData.csv_filename;
                break;
              }
              await new Promise(r => setTimeout(r, 1000));
            }
            progressBar.style.width = '100%';
            progressPercentText.innerText = '100%';
            if (csvFilename) {
              document.getElementById('download-link').innerHTML =
                `<b>Validation complete! Click the link below to download your results:</b><br>
                <a href="/download/${csvFilename}" download>Download CSV</a>`;
            } else {
              document.getElementById('download-link').innerText = 'Validation failed.';
            }
          } else {
            document.getElementById('download-link').innerText = 'Upload failed.';
          }
        }
      };
      xhr.send(formData);
    }
    </script>
    ''')

@app.route('/api/validate', methods=['POST'])
def api_validate():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(upload_path)

        # Generate a unique progress key
        progress_key = str(uuid.uuid4())
        progress_store[progress_key] = {'percent': 0, 'done': False, 'csv_filename': None}

        def run_validation():
            try:
                print(f"Starting validation for {upload_path}")
                df, csv_filename = validate_file(upload_path, progress_key, progress_key)
                progress_store[progress_key]['csv_filename'] = csv_filename
                print(f"Validation finished for {upload_path}, CSV: {csv_filename}")
            except Exception as e:
                error_csv = os.path.splitext(filename)[0] + "_validation_summary.csv"
                error_csv_path = os.path.join(EXPORTS_FOLDER, error_csv)
                with open(error_csv_path, "w", newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Error"])
                    writer.writerow([str(e)])
                progress_store[progress_key]['csv_filename'] = error_csv
                print(f"Validation error: {e}")
            finally:
                progress_store[progress_key]['percent'] = 100
                progress_store[progress_key]['done'] = True
                print(f"Validation thread complete for {upload_path}")

        # Run validation in a background thread
        thread = threading.Thread(target=run_validation)
        thread.start()

        return jsonify({'progressKey': progress_key})
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/api/progress/<progress_key>', methods=['GET'])
def get_progress(progress_key):
    prog = progress_store.get(progress_key)
    print(f"Progress check for key {progress_key}: {prog}")
    if not prog:
        return jsonify({'percent': 0, 'done': False})
    return jsonify({
        'percent': prog.get('percent', 0),
        'done': prog.get('done', False),
        'csv_filename': prog.get('csv_filename')
    })

@app.route('/download/<csv_filename>', methods=['GET'])
def download_csv(csv_filename):
    try:
        print(f"Download requested for {csv_filename}")
        return send_from_directory(app.config['EXPORTS_FOLDER'], csv_filename, as_attachment=True)
    except FileNotFoundError:
        print(f"File not found for download: {csv_filename}")
        return "File not found", 404

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 3000))
    app.run(host='0.0.0.0', port=port)